from string import ascii_uppercase
from datetime import datetime as dt
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import PatternFill, Font
from openpyxl.drawing.image import Image


def generate_subsidiary_goal_excel_file(data, qs):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Identificador'
    ws['B1'] = 'Meta'
    ws['C1'] = 'Filial'
    ws['D1'] = 'Total meta anual'
    ws['E1'] = 'Ene'
    ws['F1'] = 'Feb'
    ws['G1'] = 'Mar'
    ws['H1'] = 'Abr'
    ws['I1'] = 'May'
    ws['J1'] = 'Jun'
    ws['K1'] = 'Jul'
    ws['L1'] = 'Ago'
    ws['M1'] = 'Sep'
    ws['N1'] = 'Oct'
    ws['O1'] = 'Nov'
    ws['P1'] = 'Dic'
    ws['Q1'] = 'Ponderación'

    x = 2
    for item in data:
        ws[f'A{x}'].value = item.pk
        ws[f'B{x}'].value = item.id_goal.description
        ws[f'C{x}'].value = f'{item.id_cost_center.zone} | {item.id_cost_center}'
        ws[f'D{x}'].value = item.annual_amount_subsidiary
        ws[f'E{x}'].value = item.amount_january
        ws[f'F{x}'].value = item.amount_february
        ws[f'G{x}'].value = item.amount_march
        ws[f'H{x}'].value = item.amount_april
        ws[f'I{x}'].value = item.amount_may
        ws[f'J{x}'].value = item.amount_june
        ws[f'K{x}'].value = item.amount_july
        ws[f'L{x}'].value = item.amount_august
        ws[f'M{x}'].value = item.amount_september
        ws[f'N{x}'].value = item.amount_october
        ws[f'O{x}'].value = item.amount_november
        ws[f'P{x}'].value = item.amount_december
        ws[f'Q{x}'].value = item.ponderation
        x = x + 1
    file_name = f'formato_metas_filial_{dt.now()}'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={file_name}.xlsx'
    wb.save(response)
    return response


def generate_subsidiary_goal_execute_excel_file(data, qs):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Identificador'
    ws['B1'] = 'Meta'
    ws['C1'] = 'Filial'
    ws['D1'] = 'Ene'
    ws['E1'] = 'Feb'
    ws['F1'] = 'Mar'
    ws['G1'] = 'Abr'
    ws['H1'] = 'May'
    ws['I1'] = 'Jun'
    ws['J1'] = 'Jul'
    ws['K1'] = 'Ago'
    ws['L1'] = 'Sep'
    ws['M1'] = 'Oct'
    ws['N1'] = 'Nov'
    ws['O1'] = 'Dic'

    x = 2
    for item in data:
        ws[f'A{x}'].value = item.pk
        ws[f'B{x}'].value = item.id_goal.description
        ws[f'C{x}'].value = f'{item.id_cost_center.zone} | {item.id_cost_center}'
        ws[f'D{x}'].value = item.amount_exec_january
        ws[f'E{x}'].value = item.amount_exec_february
        ws[f'F{x}'].value = item.amount_exec_march
        ws[f'G{x}'].value = item.amount_exec_april
        ws[f'H{x}'].value = item.amount_exec_may
        ws[f'I{x}'].value = item.amount_exec_june
        ws[f'J{x}'].value = item.amount_exec_july
        ws[f'K{x}'].value = item.amount_exec_august
        ws[f'L{x}'].value = item.amount_exec_september
        ws[f'M{x}'].value = item.amount_exec_october
        ws[f'N{x}'].value = item.amount_exec_november
        ws[f'O{x}'].value = item.amount_exec_december
        x = x + 1

    file_name = f'formato_ejecucion_metas_filial_{dt.now()}'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={file_name}.xlsx'
    wb.save(response)
    return response


def load_excel_file(file):
    try:
        book = load_workbook(file, data_only=True)
        sheet_list = book.sheetnames
        first_sheet = book.get_sheet_by_name(sheet_list[0])
        number_rows = first_sheet.max_row
        return {
            'status': 'ok',
            'sheet': first_sheet,
            'number_rows': number_rows
        }
    except InvalidFileException as e:
        raise InvalidFileException(
            f'Error loading excel file: {e.__str__()}'
        )


def generate_report_by_subsidiary(request, data, months_labels, total_rating):
    FONT = Font(color="FFFFFF")
    FILL = PatternFill(start_color="017224", end_color="017224", fill_type="solid")

    def _set_fomart_value(format):
        if format == 'Porcentaje':
            return '0.00%'
        elif format == 'Moneda':
            return '#,##0.00'
        else:
            return '#,##0'

    alphabet = list(ascii_uppercase)
    alphabet.remove('A')
    alphabet.remove('B')
    alphabet.remove('C')
    alphabet.remove('D')

    wb = Workbook()
    ws = wb.active
    img = Image('static/img/coop.png')
    img.anchor = 'A1'
    ws.add_image(img)
    ws['C1'].value = "REPORTE METAS"
    ws['C1'].font = Font(size="26")
    ws['C2'].value = f"Fecha: {dt.today().strftime('%d-%m-%Y')}"
    ws['C3'].value = f"Generado por: {request.user.first_name}"

    ws['A5'] = 'Filial'
    ws['A5'].fill = FILL
    ws['A5'].font = FONT

    ws['B5'] = 'Descripción'
    ws['B5'].fill = FILL
    ws['B5'].font = FONT

    ws['C5'] = 'Meta Anual'
    ws['C5'].fill = FILL
    ws['C5'].font = FONT

    ws['D5'] = 'Meta Al Mes'
    ws['D5'].fill = FILL
    ws['D5'].font = FONT

    for index, value in enumerate(months_labels, start=0):
        ws[f'{alphabet[index]}5'] = value
        ws[f'{alphabet[index]}5'].fill = FILL
        ws[f'{alphabet[index]}5'].font = FONT
        if index == len(months_labels) - 1:
            accumulated_execution_letter = alphabet[index + 1]
            ponderation_letter = alphabet[index + 2]
            percentage_letter = alphabet[index + 3]
            rating_letter = alphabet[index + 4]
            ws[f'{alphabet[index + 1]}5'] = "Ejecución Acumulada"
            ws[f'{alphabet[index + 1]}5'].fill = FILL
            ws[f'{alphabet[index + 1]}5'].font = FONT
            ws[f'{alphabet[index + 2]}5'] = "Ponderación"
            ws[f'{alphabet[index + 2]}5'].fill = FILL
            ws[f'{alphabet[index + 2]}5'].font = FONT
            ws[f'{alphabet[index + 3]}5'] = "Porcentaje"
            ws[f'{alphabet[index + 3]}5'].fill = FILL
            ws[f'{alphabet[index + 3]}5'].font = FONT
            ws[f'{alphabet[index + 4]}5'] = "Calificación (Total: {:.2f})".format(total_rating) # NOQA
            ws[f'{alphabet[index + 4]}5'].fill = FILL
            ws[f'{alphabet[index + 4]}5'].font = FONT

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 20
    ws.column_dimensions['M'].width = 20
    ws.column_dimensions['N'].width = 20
    ws.column_dimensions['O'].width = 20
    ws.column_dimensions['P'].width = 20
    ws.column_dimensions['Q'].width = 20
    ws.column_dimensions['T'].width = 20

    for counter, item in enumerate(data, start=6):
        i_format = _set_fomart_value(item.get('Formato'))
        ws[f'A{counter}'].value = item.get('DescAgencia')
        ws[f'B{counter}'].value = item.get('Nivel2')

        ws[f'C{counter}'].value = float(item.get('MetaAnual'))
        ws[f'C{counter}'].number_format = i_format
        ws[f'D{counter}'].value = float(item.get('MetaAlMes'))
        ws[f'D{counter}'].number_format = i_format

        if 'Enero' in months_labels:
            ws[f'E{counter}'].value = float(item.get('EneroEjecucion'))
            ws[f'E{counter}'].number_format = i_format
        if 'Febrero' in months_labels:
            ws[f'F{counter}'].value = float(item.get('FebreroEjecucion'))
            ws[f'F{counter}'].number_format = i_format
        if 'Marzo' in months_labels:
            ws[f'G{counter}'].value = float(item.get('MarzoEjecucion'))
            ws[f'G{counter}'].number_format = i_format
        if 'Abril' in months_labels:
            ws[f'H{counter}'].value = float(item.get('AbrilEjecucion'))
            ws[f'H{counter}'].number_format = i_format
        if 'Mayo' in months_labels:
            ws[f'I{counter}'].value = float(item.get('MayoEjecucion'))
            ws[f'I{counter}'].number_format = i_format
        if 'Junio' in months_labels:
            ws[f'J{counter}'].value = float(item.get('JunioEjecucion'))
            ws[f'J{counter}'].number_format = i_format
        if 'Julio' in months_labels:
            ws[f'K{counter}'].value = float(item.get('JulioEjecucion'))
            ws[f'K{counter}'].number_format = i_format
        if 'Agosto' in months_labels:
            ws[f'L{counter}'].value = float(item.get('AgostoEjecucion'))
            ws[f'L{counter}'].number_format = i_format
        if 'Septiembre' in months_labels:
            ws[f'M{counter}'].value = float(item.get('SeptiembreEjecucion'))
            ws[f'M{counter}'].number_format = i_format
        if 'Octubre' in months_labels:
            ws[f'N{counter}'].value = float(item.get('OctubreEjecucion'))
            ws[f'N{counter}'].number_format = i_format
        if 'Noviembre' in months_labels:
            ws[f'O{counter}'].value = float(item.get('NoviembreEjecucion'))
            ws[f'O{counter}'].number_format = i_format
        if 'Diciembre' in months_labels:
            ws[f'P{counter}'].value = float(item.get('DiciembreEjecucion'))
            ws[f'P{counter}'].number_format = i_format
        ws[f'{accumulated_execution_letter}{counter}'].value = float(item.get("EjecucionAcumulada")) # NOQA
        ws[f'{accumulated_execution_letter}{counter}'].number_format = i_format

        ws[f'{ponderation_letter}{counter}'].value = float(item.get('Ponderacion'))
        ws[f'{ponderation_letter}{counter}'].number_format = _set_fomart_value('Cantidad')

        ws[f'{percentage_letter}{counter}'].value = float(item.get("Porcentaje"))
        ws[f'{percentage_letter}{counter}'].number_format = _set_fomart_value('Porcentaje')

        ws[f'{rating_letter}{counter}'].value = float(item.get("Calificacion"))
        ws[f'{rating_letter}{counter}'].number_format = _set_fomart_value('Moneda')

    file_name = f'reporte_{dt.now()}'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={file_name}.xlsx'
    wb.save(response)
    return response
