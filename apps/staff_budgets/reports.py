import os
import datetime as dt
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font, Border, Side, PatternFill

from ppto_safa.constants import STAFF_POSITIONS


def create_excel_report(qs):
    wb = Workbook()
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    dest_filename = os.path.join(BASE_DIR, "empty_book.xlsx")
    ws1 = wb.active
    ws1.title = "PPTO. PERSONAL"

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    today = str(dt.date.today())
    row = 3
    ws1.cell(column=1, row=row).fill = PatternFill(fgColor="2FC300", fill_type="solid")
    ws1.cell(column=2, row=row).fill = PatternFill(fgColor="2FC300", fill_type="solid")
    ws1.cell(column=3, row=row).fill = PatternFill(fgColor="2FC300", fill_type="solid")
    ws1.cell(column=4, row=row).fill = PatternFill(fgColor="2FC300", fill_type="solid")
    ws1.cell(column=5, row=row).fill = PatternFill(fgColor="2FC300", fill_type="solid")
    ws1.cell(column=6, row=row).fill = PatternFill(fgColor="2FC300", fill_type="solid")

    cost_center = period = ''
    if len(qs) > 0:
        cost_center = qs[0]["codcentrocosto__desccentrocosto"]
        period = qs[0]["periodo__descperiodo"]

    _ = ws1.cell(
        column=3,
        row=2,
        value=f'Presupuesto Personal de {cost_center}',
    )
    _.font = Font(bold=True, size="26")

    _ = ws1.cell(column=1, row=row, value="{0}".format("Tipo Personal"))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=2, row=row, value="{0}".format("Puesto"))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=3, row=row, value="{0}".format("Cantidad"))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=4, row=row, value="{0}".format("Mes Inicio"))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=5, row=row, value="{0}".format("Mes Fin"))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=6, row=row, value="{0}".format("Justificación"))
    _.font = Font(bold=True)
    _.border = border

    for i in qs:
        row += 1
        _ = ws1.cell(
            column=1, row=row, value=f"{STAFF_POSITIONS.get(i['tipo'])}"
        ).border = border
        _ = ws1.cell(
            column=2, row=row, value=f'{ i["codpuesto__descpuesto"] }'
        ).border = border
        _ = ws1.cell(
            column=3, row=row, value=f'{ i["cantidad"] }'
        ).border = border
        _ = ws1.cell(column=4, row=row, value=f'{i["mes"]}').border = border
        _ = ws1.cell(
            column=5, row=row, value=f'{i["mesfin"] if i["mesfin"] else ""}'
        ).border = border
        _ = ws1.cell(
            column=6, row=row, value=f'{i["justificacion"]}').border = border
    file_name = (
        f'PPTO. PERSONAL {period}'
        f'{cost_center} ({today.replace("-","")})'
    )
    ws1.column_dimensions["A"].width = 15
    ws1.column_dimensions["B"].width = 30
    ws1.column_dimensions["C"].width = 10
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 12
    ws1.column_dimensions["F"].width = 100
    wb.save(filename=dest_filename)
    stream = save_virtual_workbook(wb)
    response = HttpResponse(stream, content_type="text/xlsx")
    response["Content-Disposition"] = f"attachment; filename={file_name}.xlsx"
    return response
