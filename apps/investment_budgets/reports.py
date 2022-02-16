import os
import datetime as dt
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font, Border, Side, PatternFill


def create_excel_report(qs):
    wb = Workbook()
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    dest_filename = os.path.join(BASE_DIR, "empty_book.xlsx")
    ws1 = wb.active
    ws1.title = "PPTO. INVERSION"

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    row = 3
    ws1.cell(column=1, row=row).fill = PatternFill(fgColor="2FC300", fill_type="solid")
    ws1.cell(column=2, row=row).fill = PatternFill(fgColor="2FC300", fill_type="solid")
    ws1.cell(column=3, row=row).fill = PatternFill(fgColor="2FC300", fill_type="solid")
    ws1.cell(column=4, row=row).fill = PatternFill(fgColor="2FC300", fill_type="solid")
    ws1.cell(column=5, row=row).fill = PatternFill(fgColor="2FC300", fill_type="solid")
    ws1.cell(column=6, row=row).fill = PatternFill(fgColor="2FC300", fill_type="solid")
    ws1.cell(column=7, row=row).fill = PatternFill(fgColor="2FC300", fill_type="solid")

    _ = ws1.cell(
        column=3,
        row=1,
        value="Presupuesto de Inversión:"
    )
    _.font = Font(bold=True, size="26")

    _ = ws1.cell(column=1, row=row, value="{0}".format("Centro de Costos"))
    _.font = Font(bold=True)
    _ = ws1.cell(column=2, row=row, value="{0}".format("Cuenta Contable"))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=3, row=row, value="{0}".format("Mes"))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=4, row=row, value="{0}".format("Producto"))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=5, row=row, value="{0}".format("Cantidad"))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=6, row=row, value="{0}".format("Precio Unitario"))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=7, row=row, value="{0}".format("Total"))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=8, row=row, value="{0}".format("Justificación"))
    _.font = Font(bold=True)
    _.border = border

    for i in qs:
        row += 1
        _ = ws1.cell(
            column=1,
            row=row,
            value=f'{i["codcentrocostoxcuentacontable__codcentrocosto__desccentrocosto"]}', # NOQA
        ).border = border
        _ = ws1.cell(
            column=2,
            row=row,
            value=f'{i["codcentrocostoxcuentacontable__codcuentacontable__desccuentacontable"]}', # NOQA
        ).border = border
        _ = ws1.cell(column=3, row=row, value="{0}".format(i["mes"])).border = border
        _ = ws1.cell(
            column=4, row=row, value="{0}".format(i["descproducto"])
        ).border = border
        _ = ws1.cell(
            column=5, row=row, value="{0}".format(i["cantidad"])
        ).border = border
        _ = ws1.cell(
            column=6, row=row, value="{0}".format(float(i["valor"]))
        ).border = border
        _ = ws1.cell(
            column=7, row=row, value="{0}".format(float(i["presupuestado"]))
        ).border = border
        _ = ws1.cell(
            column=8,
            row=row,
            value=f'{i["justificacion"]}',
        ).border = border

    desc = ''
    fecha = str(dt.date.today())
    if len(qs) > 0:
        desc = qs[0]["codcentrocostoxcuentacontable__codcentrocosto__desccentrocosto"]
        desc_last = qs[-1][
            "codcentrocostoxcuentacontable__codcentrocosto__desccentrocosto"
        ]
        if desc != desc_last:
            desc = 'Todos CECO'
    file_name = (
        f'PPTO. INVERSION '
        f'{desc} ({fecha.replace("-", "")})'
    )

    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 15
    ws1.column_dimensions["C"].width = 35
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 20
    ws1.column_dimensions["F"].width = 20
    ws1.column_dimensions["G"].width = 100
    wb.save(filename=dest_filename)
    stream = save_virtual_workbook(wb)
    response = HttpResponse(stream, content_type="text/xlsx")
    response["Content-Disposition"] = f"attachment; filename={file_name}.xlsx"
    return response
