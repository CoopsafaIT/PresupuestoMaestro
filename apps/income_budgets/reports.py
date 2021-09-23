from django.http import HttpResponse

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def generate_excel_file_format(qs):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Centro de costo por cuenta contable'
    ws['B1'] = 'Centro de costo'
    ws['C1'] = 'Cuenta contable'
    ws['D1'] = 'Presupuesto'
    ws['E1'] = 'Periodo'
    x = 2
    for item in qs:
        ws[f'A{x}'].value = item.codcentrocostoxcuentacontable_new.codcentrocostoxcuentacontable # NOQA
        ws[f'B{x}'].value = item.codcentrocostoxcuentacontable_new.codcentrocosto.desccentrocosto # NOQA
        ws[f'C{x}'].value = item.codcentrocostoxcuentacontable_new.codcuentacontable.desccuentacontable # NOQA
        ws[f'D{x}'].value = 0
        ws[f'E{x}'].value = item.periodo.descperiodo
        x = x + 1

    file_name = 'format_income_budget'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={file_name}.xlsx'
    wb.save(response)
    return response


def load_excel_file(file):
    try:
        book = load_workbook(file)
        sheet_list = book.sheetnames
        first_sheet = book.get_sheet_by_name(sheet_list[0])
        number_rows = first_sheet.max_row
        return {
            'status': 'ok',
            'sheet': first_sheet,
            'number_rows': number_rows
        }
    except InvalidFileException as e:
        raise InvalidFileException(
            f'Error loading excel file: {e.__str__()}'
        )
