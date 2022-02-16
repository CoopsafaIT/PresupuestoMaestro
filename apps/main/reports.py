import os
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
import openpyxl


def budget_execution_report(data):
    wb = Workbook()
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    dest_filename = os.path.join(BASE_DIR, 'empty_book.xlsx')

    ws1 = wb.active
    row2 = 3
    row = 4
    ws1.cell(row=3, column=3).fill = PatternFill(
        fgColor='2FC300', fill_type='solid'
    )
    ws1.cell(row=3, column=4).fill = PatternFill(
        fgColor='279B02', fill_type='solid'
    )
    ws1.cell(row=3, column=5).fill = PatternFill(
        fgColor='279B02', fill_type='solid'
    )
    ws1.cell(row=3, column=6).fill = PatternFill(
        fgColor='279B02', fill_type='solid'
    )
    ws1.cell(row=3, column=7).fill = PatternFill(
        fgColor='2FC300', fill_type='solid'
    )
    ws1.cell(row=3, column=8).fill = PatternFill(
        fgColor='2FC300', fill_type='solid'
    )
    ws1.cell(row=3, column=9).fill = PatternFill(
        fgColor='2FC300', fill_type='solid'
    )
    ws1.cell(column=1, row=row).fill = PatternFill(
        fgColor='9C9C9C', fill_type='solid'
    )
    ws1.cell(column=2, row=row).fill = PatternFill(
        fgColor='9C9C9C', fill_type='solid'
    )
    ws1.cell(column=3, row=row).fill = PatternFill(
        fgColor='9C9C9C', fill_type='solid'
    )
    ws1.cell(column=4, row=row).fill = PatternFill(
        fgColor='9C9C9C', fill_type='solid'
    )
    ws1.cell(column=5, row=row).fill = PatternFill(
        fgColor='9C9C9C', fill_type='solid'
    )
    ws1.cell(column=6, row=row).fill = PatternFill(
        fgColor='9C9C9C', fill_type='solid'
    )
    ws1.cell(column=7, row=row).fill = PatternFill(
        fgColor='9C9C9C', fill_type='solid'
    )
    ws1.cell(column=8, row=row).fill = PatternFill(
        fgColor='9C9C9C', fill_type='solid'
    )
    ws1.cell(column=9, row=row).fill = PatternFill(
        fgColor='9C9C9C', fill_type='solid'
    )

    img = openpyxl.drawing.image.Image('static/img/coop.png')
    img.anchor = 'C2'
    ws1.add_image(img)

    ws1.title = "Ejecucion presupuestaria"
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    alignment = Alignment(horizontal='center',
                          vertical='center', wrap_text=True)

    _ = ws1.cell(column=3, row=row2, value="{0}".format('Anual'))
    _.font = Font(bold=True)
    _.alignment = alignment
    _ = ws1.cell(column=5, row=row2, value="{0}".format('Mensual'))
    _.font = Font(bold=True)
    _ = ws1.cell(column=8, row=row2, value="{0}".format('Acumulado'))
    _.font = Font(bold=True)
    _ = ws1.cell(column=1, row=row, value="{0}".format('Centro de costos'))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(
        column=4, row=2, value="{0}".format('REPORTE DE EJECUCIÓN PRESUPUESTARIA')
    )
    _.font = Font(bold=True, size="26")
    _ = ws1.cell(column=2, row=row, value="{0}".format('Cuenta Contable'))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=3, row=row, value="{0}".format('Presupuesto Anual'))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=4, row=row, value="{0}".format('Presupuesto Mensual'))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=5, row=row, value="{0}".format('Ejecutado Mensual'))
    _.font = Font(bold=True)
    _.border = border

    _ = ws1.cell(column=6, row=row, value="{0}".format('Diferencia Mensual'))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=7, row=row, value="{0}".format('Presupuesto Acumulado'))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=8, row=row, value="{0}".format('Ejecutado Acumulado'))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=9, row=row, value="{0}".format('Diferencia Acumulado'))
    _.font = Font(bold=True)
    _.border = border

    for x in data:
        row += 1
        _ = ws1.cell(column=1, row=row, value="{0}".format(
            x['DescCentroCosto'])).border = border
        _ = ws1.cell(column=2, row=row, value="{0}".format(
            x['DescCuentaContable'])).border = border
        _ = ws1.cell(column=3, row=row, value=float(
            "{0:.2f}".format(x['PresupuestoAnual']))).border = border
        _ = ws1.cell(column=3, row=row, value=float("{0:.2f}".format(
            x['PresupuestoAnual']))).number_format = '#,##0.00'
        _ = ws1.cell(column=4, row=row, value=float(
            "{0:.2f}".format(x['PresupuestoMensual']))).border = border
        _ = ws1.cell(column=4, row=row, value=float("{0:.2f}".format(
            x['PresupuestoMensual']))).number_format = '#,##0.00'
        _ = ws1.cell(column=5, row=row, value=float(
            "{0:.2f}".format(x['EjecutadoMensual']))).border = border
        _ = ws1.cell(column=5, row=row, value=float("{0:.2f}".format(
            x['EjecutadoMensual']))).number_format = '#,##0.00'
        _ = ws1.cell(column=6, row=row, value=float(
            "{0:.2f}".format(x['DiferenciasMensual']))).border = border
        _ = ws1.cell(column=6, row=row, value=float("{0:.2f}".format(
            x['DiferenciasMensual']))).number_format = '#,##0.00'
        _ = ws1.cell(column=7, row=row, value=float(
            "{0:.2f}".format(x['PresupuestoAcumulado']))).border = border
        _ = ws1.cell(column=7, row=row, value=float("{0:.2f}".format(
            x['PresupuestoAcumulado']))).number_format = '#,##0.00'
        _ = ws1.cell(column=8, row=row, value=float(
            "{0:.2f}".format(x['EjecutadoAcumulado']))).border = border
        _ = ws1.cell(column=8, row=row, value=float("{0:.2f}".format(
            x['EjecutadoAcumulado']))).number_format = '#,##0.00'
        _ = ws1.cell(column=9, row=row, value=float(
            "{0:.2f}".format(x['DiferenciasAcumulado']))).border = border
        _ = ws1.cell(column=9, row=row, value=float("{0:.2f}".format(
            x['DiferenciasAcumulado']))).number_format = '#,##0.00'

    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 70
    ws1.column_dimensions['C'].width = 25
    ws1.column_dimensions['D'].width = 25
    ws1.column_dimensions['E'].width = 25
    ws1.column_dimensions['F'].width = 25
    ws1.column_dimensions['G'].width = 25
    ws1.column_dimensions['H'].width = 25
    ws1.column_dimensions['I'].width = 25
    ws1.column_dimensions['J'].width = 25

    wb.save(filename=dest_filename)
    stream = save_virtual_workbook(wb)
    response = HttpResponse(stream, content_type='text/xlsx')
    nombre = "REPORTE DE EJECUCION PRESUPUESTARIA.xlsx"
    response['Content-Disposition'] = 'attachment; filename={0}'.format(nombre)
    return response
