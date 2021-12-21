import json
import os
import datetime as dt
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.core.exceptions import PermissionDenied
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.db.models import F
from django.core.exceptions import ValidationError

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font, Border, Side

from apps.main.models import (
    Centroscosto,
    Centrocostoxcuentacontable,
    Criterios,
    Historicotraslados,
    Periodo,
    Proyectos,
    Presupuestos,
    Presupuestocostos,
    Presupuestoindirecto,
    ResponsablesPorCentrosCostos,
    Tipopresupuesto,
)
from ppto_safa.utils import execute_sql_query
from utils.constants import MONTH


@login_required()
def budget_register(request):
    if (
        not request.user.has_perm('ppto_gastos.puede_ingresar_ppto_gastos_todos') and
        not request.user.has_perm('ppto_gastos.puede_ingresar_ppto_gastos')
    ):
        raise PermissionDenied

    def _get_ceco():
        if not request.user.has_perm('ppto_gastos.puede_ingresar_ppto_gastos_todos'):
            ceco_assigned = ResponsablesPorCentrosCostos.objects.filter(
                CodUser=request.user.pk, Estado=True
            ).values_list('CodCentroCosto', flat=True)
            return Centroscosto.objects.filter(habilitado=True).filter(
                pk__in=list(ceco_assigned)
            )
        else:
            return Centroscosto.objects.filter(habilitado=True)

    def validate_aprrove_method():
        period = request.POST.get('period', None)
        cost_center = request.POST.get('cost_center', None)
        if period is None or cost_center is None:
            raise ValidationError(
                'Periodo y Centro de costo son valores obligatorios'
            )

    if request.is_ajax():
        try:
            cost_center = request.POST.get('cost_center', '')
            period = request.POST.get('period', '')
            project = request.POST.get('project', '')
            justification = request.POST.get('justification', '')
            total_amount = float(request.POST.get('total_amount', 0).replace(',', ''))
            total_per_month = total_amount / 12
            criteria = Criterios.objects.get(pk=request.POST.get('criteria', ''))

            budget_update = Presupuestos.objects.get(
                codpresupuesto=request.POST.get('id_budget', '')
            )
            budget_update.montooriginal = total_amount
            budget_update.enero = total_per_month
            budget_update.febrero = total_per_month
            budget_update.marzo = total_per_month
            budget_update.abril = total_per_month
            budget_update.mayo = total_per_month
            budget_update.junio = total_per_month
            budget_update.julio = total_per_month
            budget_update.agosto = total_per_month
            budget_update.septiembre = total_per_month
            budget_update.octubre = total_per_month
            budget_update.noviembre = total_per_month
            budget_update.diciembre = total_per_month
            budget_update.diciembre = total_per_month
            budget_update.criterio = criteria
            budget_update.justificacion = justification
            budget_update.estado = 1
            budget_update.save()

            query = (
                f'SELECT SUM(CASE WHEN Estado=1 THEN 1 ELSE 0 END) as Presupuestada, '
                f'COUNT(*) as Totales FROM [PresupuestoSafa].[dbo].[Presupuestos] A '
                f'INNER JOIN PresupuestoSafa.dbo.Periodo B on A.CodPeriodo = B.CodPeriodo '
                f'INNER JOIN PresupuestoSafa.dbo.CentroCostoXCuentaContable  C '
                f'on A.CodCentrosCostoXCuentaContable = C.CodCentroCostoXCuentaContable '
                f'Where B.CodPeriodo = {period} and C.CodCentroCosto ={cost_center} '
                f'and A.CodProyecto={project} and  CodTipoPresupuesto=1'
            )
            result_query = execute_sql_query(query)
            if result_query.get('status') == 'error':
                raise Exception(f'Error while executing query: {result_query.get("data")}')

            data = Presupuestos.objects.extra({
                'enero': 'CAST(CAST(enero as decimal (17,2)) as nvarchar(100))',
                'febrero': 'CAST(CAST(febrero as decimal (17,2)) as nvarchar(100))',
                'marzo': 'CAST(CAST(marzo as decimal (17,2)) as nvarchar(100))',
                'abril': 'CAST(CAST(abril as decimal (17,2)) as nvarchar(100))',
                'mayo': 'CAST(CAST(mayo as decimal (17,2)) as nvarchar(100))',
                'junio': 'CAST(CAST(junio as decimal (17,2)) as nvarchar(100))',
                'julio': 'CAST(CAST(julio as decimal (17,2)) as nvarchar(100))',
                'agosto': 'CAST(CAST(agosto as decimal (17,2)) as nvarchar(100))',
                'septiembre': 'CAST(CAST(septiembre as decimal (17,2)) as nvarchar(100))',
                'octubre': 'CAST(CAST(octubre as decimal (17,2)) as nvarchar(100))',
                'noviembre': 'CAST(CAST(noviembre as decimal (17,2)) as nvarchar(100))',
                'diciembre': 'CAST(CAST(diciembre as decimal (17,2)) as nvarchar(100))',
            }).values(
                'enero', 'febrero', 'marzo', 'abril',
                'mayo', 'junio', 'julio', 'agosto',
                'septiembre', 'octubre', 'noviembre', 'diciembre',
                'estado', 'pk'
            ).filter(
                codpresupuesto=request.POST.get('id_budget', '')
            )

            ctx = {
                'data': list(data),
                'result': result_query.get("data")
            }
            return HttpResponse(json.dumps(ctx), content_type='application/json')
        except Exception as e:
            response = JsonResponse({
                'message': f'{e.__str__()}'
            })
            response.status_code = 500
            return response

    if request.method == 'POST':
        if request.POST.get('method') == 'create-budget':
            try:
                cost_center = request.POST.get('cost_center', '')
                period = request.POST.get('period', '')
                project = request.POST.get('project', '')
                request.session['cost_center'] = cost_center
                request.session['period'] = period
                request.session['project'] = project

                qs = Presupuestos.objects.filter(
                    aprobadogerencia__isnull=True,
                    codperiodo=period,
                    codtipopresupuesto=1,
                    codcentroscostoxcuentacontable__codcentrocosto=cost_center
                )
                qs_requested = qs.filter(etapasolicitado__isnull=True).count()
                qs_approved_management = qs.filter(aprobadogerencia__isnull=True).count()
                qs_approved_directors = qs.filter(aprobadojuntadirectiva__isnull=True).count() # NOQA
                qs_approved_assembly = qs.filter(aprobadoasamblea__isnull=True).count()
                accounts = Centrocostoxcuentacontable.objects.filter(
                    codcentrocosto=cost_center,
                    codcuentacontable__codtipocuenta=1,
                    habilitado=True
                )
                for account in accounts:
                    qs_cost_center_per_account = Centrocostoxcuentacontable.objects.get(
                        codcentrocosto=cost_center,
                        codcuentacontable=account.codcuentacontable.pk,
                        habilitado=True
                    )
                    if not Presupuestos.objects.filter(
                        codperiodo=period,
                        codproyecto=project,
                        codtipopresupuesto=1,
                        codcentroscostoxcuentacontable=qs_cost_center_per_account.pk
                    ).exists():
                        qs_period = Periodo.objects.get(codperiodo=period)
                        qs_type_budget = Tipopresupuesto.objects.get(codtipopresupuesto=1)
                        qs_project = Proyectos.objects.get(codproyecto=project)

                        new_record = Presupuestos()
                        new_record.codperiodo = qs_period
                        new_record.codtipopresupuesto = qs_type_budget
                        new_record.codproyecto = qs_project
                        new_record.usuariocreacion = request.user
                        new_record.codcentroscostoxcuentacontable = qs_cost_center_per_account # NOQA
                        new_record.enero = 0
                        new_record.febrero = 0
                        new_record.marzo = 0
                        new_record.abril = 0
                        new_record.mayo = 0
                        new_record.junio = 0
                        new_record.julio = 0
                        new_record.agosto = 0
                        new_record.septiembre = 0
                        new_record.octubre = 0
                        new_record.noviembre = 0
                        new_record.diciembre = 0
                        new_record.estado = 0
                        new_record.saldonoviembre = 0
                        new_record.saldodiciembre = 0
                        new_record.mesproyeccion = 11
                        new_record.save()
            except (
                Periodo.DoesNotExist(),
                Tipopresupuesto.DoesNotExist(),
                Proyectos.DoesNotExist(),
                Centrocostoxcuentacontable.DoesNotExist()
            ) as e:
                messages.warning(
                    request,
                    f'Execption catched: {e.__str__() }'
                )
                return redirect('budget_register')
            except Exception as e:
                messages.error(
                    request,
                    f'Execption catched: {e.__str__() }'
                )
                return redirect('budget_register')
            else:
                qs_budget = Presupuestos.objects.filter(
                    codcentroscostoxcuentacontable__codcentrocosto=cost_center,
                    codperiodo=period,
                    codtipopresupuesto=1,
                    codproyecto=project,
                    codcentroscostoxcuentacontable__habilitado=1
                )
                month_projection = MONTH.get(int(qs_budget.first().mesproyeccion))

                qs_total_budget = qs_budget.extra(
                    {'totalpresupuestado': 'SUM(totalpresupuestado)'}
                ).values('totalpresupuestado')
                qs_total_budget = qs_total_budget[0]['totalpresupuestado']

                periods = Periodo.objects.filter(habilitado=True)
                criteria = Criterios.objects.all()
                ctx = {
                    'periods': periods,
                    'cost_centers': _get_ceco(),
                    'criteria': criteria,
                    'month_projection': month_projection,
                    'qs_total_budget': qs_total_budget,
                    'qs_budget': qs_budget,
                    'qs_requested': qs_requested,
                    'qs_approved_management': qs_approved_management,
                    'qs_approved_directors': qs_approved_directors,
                    'qs_approved_assembly': qs_approved_assembly,
                }
            finally:
                return render(request, 'budget_register.html', ctx)

        elif request.POST.get('method') == 'save-requested':
            try:
                validate_aprrove_method()
                period = request.POST.get('period', None)
                cost_center = request.POST.get('cost_center', None)

                qs = Presupuestos.objects.filter(
                    etapasolicitado__isnull=True,
                    codperiodo=period,
                    codtipopresupuesto=1,
                    codcentroscostoxcuentacontable__codcentrocosto=cost_center
                )
                if qs.count() > 0:
                    Presupuestos.objects.filter(
                        etapasolicitado__isnull=True,
                        codperiodo=period,
                        codtipopresupuesto=1,
                        codcentroscostoxcuentacontable__codcentrocosto=cost_center
                    ).update(
                        etapasolicitado=F('montooriginal')
                    )
                messages.success(
                    request,
                    'Presupuesto guardado como: solicitado!'
                )
            except ValidationError as e:
                messages.warning(request, f'{ e.__str__() }')
            except Exception as e:
                messages.error(request, f'{ e.__str__() }')

        elif request.POST.get('method') == 'save-approved-management':
            try:
                validate_aprrove_method()
                period = request.POST.get('period', None)
                cost_center = request.POST.get('cost_center', None)

                qs = Presupuestos.objects.filter(
                    aprobadoasamblea__isnull=True,
                    codperiodo=period,
                    codtipopresupuesto=1,
                    codcentroscostoxcuentacontable__codcentrocosto=cost_center
                )
                if qs.count() > 0:
                    Presupuestos.objects.filter(
                        aprobadoasamblea__isnull=True,
                        codperiodo=period,
                        codtipopresupuesto=1,
                        codcentroscostoxcuentacontable__codcentrocosto=cost_center
                    ).update(
                        aprobadoasamblea=F('montooriginal')
                    )
                messages.success(
                    request,
                    'Presupuesto guardado como: aprobado por Gerencia'
                )
            except ValidationError as e:
                messages.warning(request, f'{ e.__str__() }')
            except Exception as e:
                messages.error(request, f'{ e.__str__() }')

        elif request.POST.get('method') == 'save-approved-directors':
            try:
                validate_aprrove_method()
                period = request.POST.get('period', None)
                cost_center = request.POST.get('cost_center', None)

                qs = Presupuestos.objects.filter(
                    aprobadojuntadirectiva__isnull=True,
                    codperiodo=period,
                    codtipopresupuesto=1,
                    codcentroscostoxcuentacontable__codcentrocosto=cost_center
                )
                if qs.count() > 0:
                    Presupuestos.objects.filter(
                        aprobadojuntadirectiva__isnull=True,
                        codperiodo=period,
                        codtipopresupuesto=1,
                        codcentroscostoxcuentacontable__codcentrocosto=cost_center
                    ).update(
                        aprobadojuntadirectiva=F('montooriginal')
                    )
                messages.success(
                    request,
                    'Presupuesto guardado como: aprobado por Junta Directiva'
                )
            except ValidationError as e:
                messages.warning(request, f'{ e.__str__() }')
            except Exception as e:
                messages.error(request, f'{ e.__str__() }')

        elif request.POST.get('method') == 'save-approved-assembly':
            try:
                validate_aprrove_method()
                period = request.POST.get('period', None)
                cost_center = request.POST.get('cost_center', None)

                qs = Presupuestos.objects.filter(
                    aprobadoasamblea__isnull=True,
                    codperiodo=period,
                    codtipopresupuesto=1,
                    codcentroscostoxcuentacontable__codcentrocosto=cost_center
                )
                if qs.count() > 0:
                    Presupuestos.objects.filter(
                        aprobadoasamblea__isnull=True,
                        codperiodo=period,
                        codtipopresupuesto=1,
                        codcentroscostoxcuentacontable__codcentrocosto=cost_center
                    ).update(
                        aprobadoasamblea=F('montooriginal')
                    )
                messages.success(
                    request,
                    'Presupuesto guardado como: aprobado por Asamblea'
                )
            except ValidationError as e:
                messages.warning(request, f'{ e.__str__() }')
            except Exception as e:
                messages.error(request, f'{ e.__str__() }')

        redirect('budget_register')

    else:
        periods = Periodo.objects.filter(habilitado=True)

        ctx = {
            'periods': periods,
            'cost_centers': _get_ceco()
        }
        return render(request, 'budget_register.html', ctx)


@login_required()
def user_create_project(request):
    def validate_create_project():
        cost_center = request.POST.get('cost_center', '')
        project_desc = request.POST.get('project_desc', '')
        date_start = request.POST.get('date_start', '') 
        date_end = request.POST.get('date_end', '')
        if(cost_center == '' or project_desc == '' or date_start == '' or date_end == ''):
            raise ValidationError(
                "Cost Center, Project Name, Date Start and Date End are required"
            )

    if request.is_ajax():
        if request.method == 'POST':
            try:
                validate_create_project()
                cost_center = request.POST.get('cost_center', '')
                project_desc = request.POST.get('project_desc', '')
                date_start = request.POST.get('date_start', '')
                date_end = request.POST.get('date_end', '')

                new_project = Proyectos()
                new_project.descproyecto = project_desc
                new_project.codcentrocosto = Centroscosto.objects.get(pk=cost_center)
                new_project.fechainicio = date_start
                new_project.fechafinal = date_end
                new_project.usuariocreacion = request.user
                new_project.save()
                response = JsonResponse({
                    'message': 'ok'
                }, status=200)
            except ValidationError as e:
                response = JsonResponse({
                    'message': f'{e.__str__()}'
                }, status=400)
            except Centroscosto.DoesNotExist as e:
                response = JsonResponse({
                    'message': f'{e.__str__()}'
                }, status=404)
            except Exception as e:
                response = JsonResponse({
                    'message': f'{e.__str__()}'
                }, status=500)
            finally:
                return response


@login_required()
def get_projects_by_cost_center(request):
    if request.is_ajax():
        cost_center = request.GET.get('cost_center_id')
        result = list(Proyectos.objects.values(
            'codproyecto', 'descproyecto'
        ).filter(codcentrocosto=cost_center))
        return HttpResponse(json.dumps(result), content_type='application/json')


@login_required()
def generate_excel_report(request, project, period, cost_center):
    period = Periodo.objects.get(codperiodo=period)
    cost_center = Centroscosto.objects.get(codcentrocosto=cost_center)
    project = Proyectos.objects.get(codproyecto=project)
    presupuesto = Presupuestos.objects.extra(
        {'totalPresupuestado': 'TotalPresupuestado'}
    ).filter(
        codcentroscostoxcuentacontable__codcentrocosto=cost_center,
        codperiodo=period,
        codtipopresupuesto=1,
        codproyecto=project
    )
    wb = Workbook()
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    dest_filename = os.path.join(BASE_DIR, 'empty_book.xlsx')
    ws1 = wb.active

    row = 1
    ws1.title = f"REPORTE PRESUPUESTO GASTOS {period.descperiodo}"
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    _ = ws1.cell(column=1, row=row, value="{0}".format('CUENTA CONTABLE'))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=2, row=row, value="{0}".format('ENERO'))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=3, row=row, value="{0}".format('FEBRERO'))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=4, row=row, value="{0}".format('MARZO'))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=5, row=row, value="{0}".format('ABRIL'))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=6, row=row, value="{0}".format('MAYO'))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=7, row=row, value="{0}".format('JUNIO'))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=8, row=row, value="{0}".format('JULIO'))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=9, row=row, value="{0}".format('AGOSTO'))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=10, row=row, value="{0}".format('SEPTIEMBRE'))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=11, row=row, value="{0}".format('OCTUBRE'))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=12, row=row, value="{0}".format('NOVIEMBRE'))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=13, row=row, value="{0}".format('DICIEMBRE'))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=14, row=row, value="{0}".format('TOTAL'))
    _.font = Font(bold=True)
    _.border = border

    for x in presupuesto:
        accounting_code = x.codcentroscostoxcuentacontable.codcuentacontable
        row += 1
        _ = ws1.cell(column=1, row=row, value="{0}".format(
            '' if accounting_code is None else accounting_code
            )).border = border
        _ = ws1.cell(column=2, row=row,  value=x.enero).border = border
        _ = ws1.cell(column=2, row=row,  value=x.enero).number_format = '#,##0.00'


        _ = ws1.cell(column=3, row=row,  value=x.febrero).border = border
        _ = ws1.cell(column=3, row=row,  value=x.febrero).number_format = '#,##0.00'


        _ = ws1.cell(column=4, row=row,  value=x.marzo).border = border
        _ = ws1.cell(column=4, row=row,  value=x.marzo).number_format = '#,##0.00'

        _ = ws1.cell(column=5, row=row,  value=x.abril).border = border
        _ = ws1.cell(column=5, row=row,  value=x.abril).number_format = '#,##0.00'

        _ = ws1.cell(column=6, row=row,  value=x.mayo).border = border
        _ = ws1.cell(column=6, row=row,  value=x.mayo).number_format = '#,##0.00'

        _ = ws1.cell(column=7, row=row,  value=x.junio).border = border
        _ = ws1.cell(column=7, row=row,  value=x.junio).number_format = '#,##0.00'

        _ = ws1.cell(column=8, row=row,  value=x.julio).border = border
        _ = ws1.cell(column=8, row=row,  value=x.julio).number_format = '#,##0.00'

        _ = ws1.cell(column=9, row=row,  value=x.agosto).border = border
        _ = ws1.cell(column=9, row=row,  value=x.agosto).number_format = '#,##0.00'

        _ = ws1.cell(column=10, row=row, value=x.septiembre).border = border
        _ = ws1.cell(column=10, row=row, value=x.septiembre).number_format = '#,##0.00'

        _ = ws1.cell(column=11, row=row, value=x.octubre).border = border
        _ = ws1.cell(column=11, row=row, value=x.octubre).number_format = '#,##0.00'

        _ = ws1.cell(column=12, row=row, value=x.noviembre).border = border
        _ = ws1.cell(column=12, row=row, value=x.noviembre).number_format = '#,##0.00'

        _ = ws1.cell(column=13, row=row, value=x.diciembre).border = border
        _ = ws1.cell(column=13, row=row, value=x.diciembre).number_format = '#,##0.00'

        _ = ws1.cell(column=14, row=row, value=x.totalPresupuestado).border = border
        _ = ws1.cell(column=14, row=row, value=x.totalPresupuestado).number_format = '#,##0.00' # NOQA

    ws1.column_dimensions['A'].width = 80
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 15
    ws1.column_dimensions['F'].width = 15
    ws1.column_dimensions['G'].width = 15
    ws1.column_dimensions['H'].width = 15
    ws1.column_dimensions['I'].width = 15
    ws1.column_dimensions['J'].width = 15
    ws1.column_dimensions['K'].width = 15
    ws1.column_dimensions['L'].width = 15
    ws1.column_dimensions['M'].width = 15
    ws1.column_dimensions['N'].width = 15
    ws1.column_dimensions['O'].width = 15
    ws1.column_dimensions['P'].width = 15
    ws1.column_dimensions['Q'].width = 15
    ws1.column_dimensions['R'].width = 15
    ws1.column_dimensions['S'].width = 15
    ws1.column_dimensions['T'].width = 15
    ws1.column_dimensions['O'].width = 15
    wb.save(filename=dest_filename)

    stream = save_virtual_workbook(wb)
    response = HttpResponse(stream, content_type='text/xlsx')
    response['Content-Disposition'] = 'attachment; filename="PPTO. GASTOS '+period.descperiodo+'.xlsx"' # NOQA
    return response


@login_required()
@permission_required(
    'ppto_gastos.puede_crear_traslados_gastos', raise_exception=True
)
def transfers_expenses(request):
    def _get_cost_qs(cost_center_origin, cost_center_destination, period):
        qs_cost_origin = Presupuestocostos.objects.extra({
            'total': 'CAST(CAST(total as decimal (17,2)) as nvarchar(100))',
        }).values(
            'codpresupuestocostos',
            'codcentrocostoxcuentacontable_new__codcuentacontable__desccuentacontable',
            'total'
        ).filter(
            codcentrocostoxcuentacontable_new__codcentrocosto=cost_center_origin,
            periodo=period
        )
        qs_cost_destination = Presupuestocostos.objects.extra({
            'total': 'CAST(CAST(total as decimal (17,2)) as nvarchar(100))',
        }).values(
            'codpresupuestocostos',
            'codcentrocostoxcuentacontable_new__codcuentacontable__desccuentacontable',
            'total'
        ).filter(
            codcentrocostoxcuentacontable_new__codcentrocosto=cost_center_destination,
            periodo=period
        )
        return {
            'qs_cost_origin': qs_cost_origin,
            'qs_cost_destination': qs_cost_destination
        }

    def _get_indirect_qs(cost_center_origin, cost_center_destination, period):
        qs_indirect_origin = Presupuestoindirecto.objects.extra({
            'total': 'CAST(CAST(total as decimal (17,2)) as nvarchar(100))',
        }).values(
            'codpresupuestoindirecto',
            'codcentrocostoxcuentacontable_new__codcuentacontable__desccuentacontable',
            'total'
        ).filter(
            codcentrocostoxcuentacontable_new__codcentrocosto=cost_center_origin,
            periodo=period
        )

        qs_indirect_destination = Presupuestoindirecto.objects.extra({
            'total': 'CAST(CAST(total as decimal (17,2)) as nvarchar(100))',
        }).values(
            'codpresupuestoindirecto',
            'codcentrocostoxcuentacontable_new__codcuentacontable__desccuentacontable',
            'total'
        ).filter(
            codcentrocostoxcuentacontable_new__codcentrocosto=cost_center_destination,
            periodo=period
        )

        return {
            'qs_indirect_origin': qs_indirect_origin,
            'qs_indirect_destination': qs_indirect_destination
        }

    def _get_expenses_projects_qs(cost_center_origin, cost_center_destination):
        projects_origin = Proyectos.objects.values(
            'codproyecto', 'descproyecto'
        ).filter(codcentrocosto=cost_center_origin)

        projects_destination = Proyectos.objects.values(
            'codproyecto', 'descproyecto'
        ).filter(codcentrocosto=cost_center_destination)
        return {
            'projects_origin': projects_origin,
            'projects_destination': projects_destination
        }

    if request.is_ajax():
        period = request.GET.get('period')
        cost_center = request.GET.get('cost_center')
        project = request.GET.get('project')
        data = Presupuestos.objects.extra({
            'montooriginal': 'CAST(CAST(montooriginal as decimal (17,2)) as nvarchar(100))'
        }).values(
            'codcentroscostoxcuentacontable__codcuentacontable',
            'pk',
            'montooriginal',
            'codcentroscostoxcuentacontable__codcuentacontable__desccuentacontable',
            'codproyecto__descproyecto'
        ).filter(
            codcentroscostoxcuentacontable__codcentrocosto=cost_center,
            codperiodo=period,
            codtipopresupuesto=1,
            codproyecto=project
        )
        return HttpResponse(
            json.dumps({'data': list(data)}), content_type='application/json'
        )

    if request.method == 'POST':
        application_date = request.POST.get('application_date')
        now = dt.datetime.now()
        cost_transfer = request.POST.get('cost_transfer')
        indirect_transfer = request.POST.get('indirect_transfer')
        expenses_transfer = request.POST.get('expenses_transfer')
        if not not expenses_transfer:
            expenses_transfer = json.loads(expenses_transfer)
            budget_origin_qs = get_object_or_404(
                Presupuestos, pk=expenses_transfer.get('originId')
            )
            budget_destination_qs = get_object_or_404(
                Presupuestos, pk=expenses_transfer.get('destinationId')
            )
            amount = Decimal(expenses_transfer.get('amount'))
            per_month = Decimal(amount / 12)
            budget_origin_qs.montooriginal = budget_origin_qs.montooriginal - amount
            budget_origin_qs.enero = budget_origin_qs.enero - per_month
            budget_origin_qs.febrero = budget_origin_qs.febrero - per_month
            budget_origin_qs.marzo = budget_origin_qs.marzo - per_month
            budget_origin_qs.abril = budget_origin_qs.abril - per_month
            budget_origin_qs.mayo = budget_origin_qs.mayo - per_month
            budget_origin_qs.junio = budget_origin_qs.junio - per_month
            budget_origin_qs.julio = budget_origin_qs.julio - per_month
            budget_origin_qs.agosto = budget_origin_qs.agosto - per_month
            budget_origin_qs.septiembre = budget_origin_qs.septiembre - per_month
            budget_origin_qs.octubre = budget_origin_qs.octubre - per_month
            budget_origin_qs.noviembre = budget_origin_qs.noviembre - per_month
            budget_origin_qs.diciembre = budget_origin_qs.diciembre - per_month
            budget_origin_qs.diciembre = budget_origin_qs.diciembre - per_month
            budget_origin_qs.fechamodificacion = now
            budget_origin_qs.usuariomodificacion = request.user.pk
            budget_origin_qs.save()

            budget_destination_qs.montooriginal = budget_destination_qs.montooriginal + amount # NOQA
            budget_destination_qs.enero = budget_destination_qs.enero + per_month
            budget_destination_qs.febrero = budget_destination_qs.febrero + per_month
            budget_destination_qs.marzo = budget_destination_qs.marzo + per_month
            budget_destination_qs.abril = budget_destination_qs.abril + per_month
            budget_destination_qs.mayo = budget_destination_qs.mayo + per_month
            budget_destination_qs.junio = budget_destination_qs.junio + per_month
            budget_destination_qs.julio = budget_destination_qs.julio + per_month
            budget_destination_qs.agosto = budget_destination_qs.agosto + per_month
            budget_destination_qs.septiembre = budget_destination_qs.septiembre + per_month
            budget_destination_qs.octubre = budget_destination_qs.octubre + per_month
            budget_destination_qs.noviembre = budget_destination_qs.noviembre + per_month
            budget_destination_qs.diciembre = budget_destination_qs.diciembre + per_month
            budget_destination_qs.diciembre = budget_destination_qs.diciembre + per_month
            budget_destination_qs.fechamodificacion = now
            budget_destination_qs.usuariomodificacion = request.user.pk
            budget_destination_qs.save()

            history = Historicotraslados()
            history.fechaaplicacion = application_date
            history.codorigengastos = budget_origin_qs
            history.coddestinogastos = budget_destination_qs
            history.montoorigengastos = amount
            history.montodestino = amount
            history.codusuario = request.user
            history.fechacreacion = now
            history.save()
            messages.success(
                request,
                (
                    f'Traslado de PPTO. de Gastos realizado con éxito. '
                    f'Monto de traslado: {amount}'
                )
            )

        if not not indirect_transfer:
            indirect_transfer = json.loads(indirect_transfer)
            budget_origin_qs = get_object_or_404(
                Presupuestoindirecto, pk=indirect_transfer.get('originId')
            )
            budget_destination_qs = get_object_or_404(
                Presupuestoindirecto, pk=indirect_transfer.get('destinationId')
            )
            amount = Decimal(indirect_transfer.get('amount'))

            budget_origin_qs.total = budget_origin_qs.total - amount
            budget_origin_qs.valor = budget_origin_qs.valor - amount
            budget_origin_qs.usuariomodificacion = request.user
            budget_origin_qs.fechamodificacion = now
            budget_origin_qs.save()

            budget_destination_qs.total = budget_destination_qs.total + amount
            budget_destination_qs.valor = budget_destination_qs.valor + amount
            budget_destination_qs.usuariomodificacion = request.user
            budget_destination_qs.fechamodificacion = now
            budget_destination_qs.save()

            history = Historicotraslados()
            history.fechaaplicacion = application_date
            history.codorigenindirecto = budget_origin_qs
            history.coddestinoindirecto = budget_destination_qs
            history.montoorigenindirecto = amount
            history.montodestino = amount
            history.codusuario = request.user
            history.fechacreacion = now
            history.save()
            messages.success(
                request,
                (
                    f'Traslado de PPTO. Indirecto realizado con éxito. '
                    f'Monto de traslado: {amount}'
                )
            )

        if not not cost_transfer:
            cost_transfer = json.loads(cost_transfer)
            budget_origin_qs = get_object_or_404(
                Presupuestocostos, pk=cost_transfer.get('originId')
            )
            budget_destination_qs = get_object_or_404(
                Presupuestocostos, pk=cost_transfer.get('destinationId')
            )
            amount = Decimal(cost_transfer.get('amount'))
            budget_origin_qs.total = budget_origin_qs.total - amount
            budget_origin_qs.valor = budget_origin_qs.valor - amount
            budget_origin_qs.usuariomodificacion = request.user
            budget_origin_qs.fechamodificacion = now
            budget_origin_qs.save()

            budget_destination_qs.total = budget_destination_qs.total + amount
            budget_destination_qs.valor = budget_destination_qs.valor + amount
            budget_destination_qs.usuariomodificacion = request.user
            budget_destination_qs.fechamodificacion = now
            budget_destination_qs.save()

            history = Historicotraslados()
            history.fechaaplicacion = application_date
            history.codorigencostos = budget_origin_qs
            history.coddestinocostos = budget_destination_qs
            history.montoorigencostos = amount
            history.montodestino = amount
            history.codusuario = request.user
            history.fechacreacion = now
            history.save()
            messages.success(
                request,
                (
                    f'Traslado de PPTO. de Costos realizado con éxito. '
                    f'Monto de traslado: {amount}'
                )
            )

    if (
        request.GET.get('period') and
        request.GET.get('cost_center_origin') and
        request.GET.get('cost_center_destination')
    ):
        cost_center_destination = request.GET.get('cost_center_destination')
        request.session['period'] = request.GET.get('period')
        request.session['cost_center_origin'] = request.GET.get('cost_center_origin')
        request.session['cost_center_destination'] = cost_center_destination

    periods = Periodo.objects.filter(habilitado=True)
    cost_centers = Centroscosto.objects.filter(habilitado=True)
    ctx = {
        'periods': periods,
        'cost_centers': cost_centers
    }

    if (
        request.session.get('period') and
        request.session.get('cost_center_origin') and
        request.session.get('cost_center_destination')
    ):
        period = request.session.get('period')
        cost_center_origin = request.session.get('cost_center_origin')
        cost_center_destination = request.session.get('cost_center_destination')
        cost = _get_cost_qs(
            cost_center_origin, cost_center_destination, period
        )
        indirect = _get_indirect_qs(
            cost_center_origin, cost_center_destination, period
        )
        expenses_project = _get_expenses_projects_qs(
            cost_center_origin, cost_center_destination
        )
        ctx['cost'] = cost
        ctx['indirect'] = indirect
        ctx['expenses_project'] = expenses_project
    return render(request, 'transfers_expenses/list.html', ctx)
