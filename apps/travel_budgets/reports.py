import os
import datetime as dt

from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils.exceptions import InvalidFileException

from utils.constants import TRAVEL_CATEGORY, ZONES, TRAVEL_TYPE


def create_excel_report(qs):
    wb = Workbook()
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    dest_filename = os.path.join(BASE_DIR, 'empty_book.xlsx')
    ws1 = wb.active
    ws1.title = 'PPTO. VIATICOS'

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    fecha = str(dt.date.today())
    row2 = 3
    ws1.cell(column=1, row=row2).fill = PatternFill(fgColor='017224', fill_type='solid')
    ws1.cell(column=2, row=row2).fill = PatternFill(fgColor='017224', fill_type='solid')
    ws1.cell(column=3, row=row2).fill = PatternFill(fgColor='017224', fill_type='solid')
    ws1.cell(column=4, row=row2).fill = PatternFill(fgColor='017224', fill_type='solid')
    ws1.cell(column=5, row=row2).fill = PatternFill(fgColor='017224', fill_type='solid')
    ws1.cell(column=6, row=row2).fill = PatternFill(fgColor='017224', fill_type='solid')
    title = qs[0]['codcentrocosto__desccentrocosto'] if len(qs) > 0 else ''
    period_desc = qs[0]["codperiodo__descperiodo"] if len(qs) > 0 else ''
    _ = ws1.cell(
        column=3,
        row=1,
        value=f"Presupuesto de Viáticos de {title}"
    )
    _.font = Font(bold=True, size="26")

    _ = ws1.cell(column=1, row=row2, value="{0}".format('TIPO VIATICO'))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=2, row=row2, value="{0}".format('FILIAL/ZONA'))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=3, row=row2, value="{0}".format('CATEGORIA'))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=4, row=row2, value="{0}".format('CANTIDAD VIAJES'))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=5, row=row2, value="{0}".format('CANTIDAD DIAS'))
    _.font = Font(bold=True)
    _.border = border
    _ = ws1.cell(column=6, row=row2, value="{0}".format('JUSTIFICACION'))
    _.font = Font(bold=True)
    _.border = border

    for i in qs:
        row2 += 1
        _ = ws1.cell(column=1, row=row2, value="{0}".format(
            TRAVEL_TYPE.get(i['tipoviatico'])
        )).border = border
        _ = ws1.cell(column=2, row=row2, value="{0}".format(
             i['filial__nombrefilial'] if i['tipoviatico'] == 1 else ZONES.get(i['zona'])
        )).border = border
        _ = ws1.cell(column=3, row=row2, value="{0}".format(
            TRAVEL_CATEGORY.get(i['categoria'])
        )).border = border
        _ = ws1.cell(column=4, row=row2, value="{:.2f}".format(
            round(i['cantidadviajes'], 2)
        )).border = border
        _ = ws1.cell(column=5, row=row2, value="{:.2f}".format(
            round(i['cantidaddias'], 2)
        )).border = border
        _ = ws1.cell(column=6, row=row2, value="{0}".format(
            i['justificacion']
        )).border = border

    file_name = (
        f'PPTO. VIATICOS {period_desc} '
        f'{title} '
        f'( {fecha.replace("-",".")} )'
    )
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 60
    ws1.column_dimensions['D'].width = 40
    ws1.column_dimensions['E'].width = 40
    ws1.column_dimensions['F'].width = 100
    wb.save(filename=dest_filename)
    stream = save_virtual_workbook(wb)
    response = HttpResponse(stream, content_type='text/xlsx')
    response['Content-Disposition'] = 'attachment; filename={0}.xlsx'.format(file_name)
    return response


def generate_excel_file_format_distribution(qs):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Identificador (No cambiar)'
    ws['B1'] = 'Periodo'
    ws['C1'] = 'Centro de costo presupuesta'
    ws['D1'] = 'Filial'
    ws['E1'] = 'Categoria'
    ws['F1'] = 'Cantidad'
    ws['G1'] = 'Días'
    ws['H1'] = 'Monto a Asignar'
    x = 2
    period = qs.first()
    for item in qs:
        ws[f'A{x}'].value = item.pk
        ws[f'B{x}'].value = item.codperiodo.descperiodo
        ws[f'C{x}'].value = item.codcentrocosto.desccentrocosto
        ws[f'D{x}'].value = item.filial.nombrefilial
        ws[f'E{x}'].value = item.categoria
        ws[f'F{x}'].value = item.cantidadviajes
        ws[f'G{x}'].value = item.cantidaddias
        ws[f'H{x}'].value = 0
        x = x + 1

    file_name = (
        f'travel_budget_distribution_format_'
        f'{period.codperiodo.descperiodo if period else ""}'
    )
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={file_name}.xlsx'
    wb.save(response)
    return response


def load_excel_file(file):
    try:
        book = load_workbook(file)
        sheet_list = book.sheetnames
        first_sheet = book.get_sheet_by_name(sheet_list[0])
        number_rows = first_sheet.max_row
        return {
            'status': 'ok',
            'sheet': first_sheet,
            'number_rows': number_rows
        }
    except InvalidFileException as e:
        raise InvalidFileException(
            f'Error loading excel file: {e.__str__()}'
        )
